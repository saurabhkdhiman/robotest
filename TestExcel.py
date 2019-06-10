import requests
from time import sleep
from json import dumps
import openpyxl


#Get the event from excel
loc = ("C:\\Users\\S Kumar\\Documents\\Staging_ClaimAPI_Sheet.xlsx")
wb = openpyxl.load_workbook(loc)
sheet = wb.active

for k in range(2,5):
    campaignId= sheet.cell(row=k, column=1).value
    phoneNumber= sheet.cell(row=k, column=2).value
